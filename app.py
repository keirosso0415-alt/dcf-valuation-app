"""
DCF Valuation App - Streamlit UI
=================================
DCFバリュエーションモデル
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from dcf_model.core import DCFModel, Assumptions, HistoricalData, Scenario
import requests
import os

# PDF生成用
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


def get_stock_price(ticker_code: str) -> tuple:
    """
    Yahoo Finance APIから株価を取得
    
    Args:
        ticker_code: 証券コード（4桁の数字）
    
    Returns:
        (株価, 会社名) または (None, エラーメッセージ)
    """
    try:
        # 日本株の場合は.Tをつける
        symbol = f"{ticker_code}.T"
        
        # Yahoo Finance APIを直接呼び出し
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
        }
        params = {
            "interval": "1d",
            "range": "5d"
        }
        
        response = requests.get(url, headers=headers, params=params, timeout=10)
        
        if response.status_code != 200:
            return None, f"APIエラー: ステータスコード {response.status_code}"
        
        data = response.json()
        
        # エラーチェック
        if "chart" not in data or "result" not in data["chart"] or not data["chart"]["result"]:
            return None, "株価データが見つかりませんでした"
        
        result = data["chart"]["result"][0]
        
        # 会社名を取得
        meta = result.get("meta", {})
        company_name = meta.get("longName") or meta.get("shortName") or meta.get("symbol", symbol)
        
        # 最新の終値を取得
        indicators = result.get("indicators", {})
        quote = indicators.get("quote", [{}])[0]
        closes = quote.get("close", [])
        
        # 最新の有効な終値を取得
        price = None
        for p in reversed(closes):
            if p is not None:
                price = float(p)
                break
        
        if price is None:
            return None, "株価データが取得できませんでした"
        
        return price, company_name
        
    except requests.exceptions.Timeout:
        return None, "タイムアウト: 接続に時間がかかっています"
    except requests.exceptions.RequestException as e:
        return None, f"通信エラー: {str(e)}"
    except Exception as e:
        return None, f"取得エラー: {str(e)}"


def get_risk_free_rate() -> tuple:
    """
    日本10年国債利回り（リスクフリーレート）を取得
    
    Returns:
        (利回り, ステータスメッセージ) または (None, エラーメッセージ)
    """
    try:
        # 日本10年国債利回り（^TNX は米国、日本は別のティッカー）
        # Yahoo Financeで日本国債は取得が難しいため、フォールバック値を使用
        # 実際の実装では財務省APIやBloombergなどを使用
        
        # 2024年時点の日本10年国債利回りの目安
        # 実務では最新値を取得するAPIを使用することを推奨
        default_rate = 0.0087  # 0.87%
        
        # 財務省の金利情報ページからスクレイピングする場合はここに実装
        # 現状はデフォルト値を返す
        
        return default_rate, "デフォルト値（最新値は財務省HPで確認）"
        
    except Exception as e:
        return None, f"取得エラー: {str(e)}"


def calculate_cost_of_debt(extracted: dict) -> tuple:
    """
    EDINETデータから負債コストを推計
    
    負債コスト ≒ 支払利息 ÷ 有利子負債
    
    Returns:
        (負債コスト, ステータス) または (None, エラーメッセージ)
    """
    try:
        # 有利子負債を計算（円単位）
        short_term = extracted.get("short_term_debt", {}).get("value") or 0
        long_term = extracted.get("long_term_debt", {}).get("value") or 0
        total_debt = short_term + long_term
        
        if total_debt <= 0:
            return None, "有利子負債データがありません"
        
        # 支払利息（EDINETから取得できる場合）
        interest_expense = extracted.get("interest_expense", {}).get("value")
        
        if interest_expense and interest_expense > 0:
            # 負債コスト = 支払利息 / 有利子負債
            cost_of_debt = interest_expense / total_debt
            return cost_of_debt, f"支払利息/有利子負債から算出"
        else:
            # 支払利息が取得できない場合はデフォルト値
            return 0.008, "デフォルト値（支払利息データなし）"
            
    except Exception as e:
        return None, f"計算エラー: {str(e)}"

# ページ設定
st.set_page_config(
    page_title="DCFバリュエーション",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Google翻訳防止用のmetaタグとlang属性
st.markdown("""
<html lang="ja">
<meta name="google" content="notranslate">
""", unsafe_allow_html=True)

# カスタムCSS
st.markdown("""
<style>
    /* カラーパレット */
    :root {
        --navy: #1e3a5f;
        --gold: #c9a962;
        --light-gray: #f5f5f5;
        --white: #ffffff;
    }

    /* 全体のフォント設定 */
    .stApp {
        font-family: 'Helvetica Neue', Arial, sans-serif;
    }

    /* ヘッダー */
    .main-header {
        background: linear-gradient(135deg, #1e3a5f 0%, #2d5a8a 100%);
        padding: 1.5rem 2rem;
        border-radius: 10px;
        margin-bottom: 1.5rem;
        color: white;
    }

    .main-header h1 {
        font-family: 'Georgia', 'Times New Roman', serif;
        margin: 0;
        font-size: 1.8rem;
        font-weight: 600;
    }

    .main-header p {
        margin: 0.5rem 0 0 0;
        opacity: 0.9;
        font-size: 0.95rem;
    }

    /* バリューカード */
    .value-card {
        background: white;
        border-radius: 10px;
        padding: 1.2rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        border-left: 4px solid #c9a962;
        height: 100%;
    }

    .value-card-title {
        font-size: 0.85rem;
        color: #666;
        margin-bottom: 0.5rem;
        font-weight: 500;
    }

    .value-card-value {
        font-family: Arial, sans-serif;
        font-size: 1.6rem;
        font-weight: 700;
        color: #1e3a5f;
        margin-bottom: 0.3rem;
    }

    .value-card-sub {
        font-size: 0.8rem;
        color: #888;
    }

    .upside-positive {
        color: #2e7d32;
    }

    .upside-negative {
        color: #c62828;
    }

    /* セクションヘッダー */
    .section-header {
        font-family: 'Georgia', 'Times New Roman', serif;
        color: #1e3a5f;
        font-size: 1.2rem;
        font-weight: 600;
        margin: 1.5rem 0 1rem 0;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #c9a962;
    }

    /* サイドバー */
    .css-1d391kg {
        background-color: #f8f9fa;
    }

    .sidebar-header {
        font-family: 'Georgia', 'Times New Roman', serif;
        color: #1e3a5f;
        font-size: 1rem;
        font-weight: 600;
        margin: 1rem 0 0.5rem 0;
        padding-bottom: 0.3rem;
        border-bottom: 1px solid #c9a962;
    }

    /* テーブルスタイル */
    .dataframe {
        font-family: Arial, sans-serif;
        font-size: 0.85rem;
    }

    /* シナリオボタン */
    .scenario-btn {
        padding: 0.5rem 1rem;
        border-radius: 5px;
        font-weight: 500;
        transition: all 0.2s;
    }

    /* Plotlyグラフのフォント */
    .js-plotly-plot .plotly text {
        font-family: Arial, sans-serif !important;
    }

    /* メトリックカスタム */
    [data-testid="stMetricValue"] {
        font-family: Arial, sans-serif;
        color: #1e3a5f;
    }

    /* 感応度分析テーブル */
    .sensitivity-table {
        font-family: Arial, sans-serif;
    }

    .sensitivity-table th {
        background-color: #1e3a5f !important;
        color: white !important;
    }
</style>
""", unsafe_allow_html=True)


def format_currency(value: float, unit: str = "百万円") -> str:
    """通貨フォーマット"""
    if abs(value) >= 1000:
        return f"{value:,.0f}{unit}"
    return f"{value:,.1f}{unit}"


def format_currency_oku(value: float) -> str:
    """通貨フォーマット（億円表示）
    
    valueは百万円単位を想定。100で割って億円に変換。
    """
    value_oku = value / 100
    if abs(value_oku) >= 100:
        return f"{value_oku:,.0f}億円"
    elif abs(value_oku) >= 10:
        return f"{value_oku:,.1f}億円"
    else:
        return f"{value_oku:,.2f}億円"


def format_stock_price(value: float) -> str:
    """株価フォーマット"""
    return f"{value:,.0f}円"


def format_percent(value: float) -> str:
    """パーセントフォーマット"""
    return f"{value:.2%}"


def create_excel_template() -> BytesIO:
    """財務データ入力用Excelテンプレートを作成（入力しやすいフォーマット）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "財務データ"
    
    # スタイル定義
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    section_font = Font(bold=True, size=11, color="1E3A5F")
    section_fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")
    input_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")  # 薄黄色
    normal_font = Font(size=10)
    unit_font = Font(size=9, color="666666")
    thin_border = Border(
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # ヘッダー行
    headers = ["項目", "値", "単位", "説明"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # データ定義（セクション分け）
    template_rows = [
        # セクション: 基本情報
        {"section": "基本情報"},
        {"item": "会社名", "value": "サンプル株式会社", "unit": "—", "desc": "分析対象の会社名"},
        {"item": "基準年度", "value": 2025, "unit": "年", "desc": "財務データの基準年度"},
        
        # セクション: P/L項目
        {"section": "損益計算書（P/L）"},
        {"item": "売上収益", "value": 500000, "unit": "百万円", "desc": "基準年の売上収益"},
        {"item": "前期売上収益", "value": 450000, "unit": "百万円", "desc": "前期の売上収益（成長率計算用）"},
        {"item": "売上原価", "value": 312500, "unit": "百万円", "desc": "売上原価（原価率計算用）"},
        {"item": "営業利益", "value": 47500, "unit": "百万円", "desc": "営業利益（販管費率計算用）"},
        {"item": "減価償却費", "value": 25000, "unit": "百万円", "desc": "減価償却費（EBITDA計算用）"},
        {"item": "当期純利益", "value": 30000, "unit": "百万円", "desc": "親会社株主に帰属する当期純利益（PER計算用）"},
        
        # セクション: B/S項目
        {"section": "貸借対照表（B/S）"},
        {"item": "現金及び現金同等物", "value": 50000, "unit": "百万円", "desc": "現金・預金等"},
        {"item": "売上債権", "value": 80000, "unit": "百万円", "desc": "売掛金・受取手形等"},
        {"item": "棚卸資産", "value": 40000, "unit": "百万円", "desc": "商品・製品・原材料等"},
        {"item": "有形固定資産", "value": 150000, "unit": "百万円", "desc": "土地・建物・設備等"},
        {"item": "短期借入金", "value": 30000, "unit": "百万円", "desc": "1年内返済予定の借入金"},
        {"item": "長期借入金", "value": 100000, "unit": "百万円", "desc": "返済期限1年超の借入金・社債"},
        
        # セクション: 株式情報
        {"section": "株式情報"},
        {"item": "発行済株式数（百万株）", "value": 100.0, "unit": "百万株", "desc": "希薄化後発行済株式数"},
        {"item": "現在株価（円）", "value": 3000, "unit": "円", "desc": "現在の株価（参考値）"},
    ]
    
    # データ書き込み
    row = 2
    for data in template_rows:
        if "section" in data:
            # セクションヘッダー
            cell = ws.cell(row=row, column=1, value=data["section"])
            cell.font = section_font
            cell.fill = section_fill
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = section_fill
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        else:
            # データ行
            ws.cell(row=row, column=1, value=data["item"]).font = normal_font
            
            # 入力セル（薄黄色で強調）
            value_cell = ws.cell(row=row, column=2, value=data["value"])
            value_cell.fill = input_fill
            value_cell.alignment = Alignment(horizontal='right')
            
            # 数値フォーマット（カンマ区切り）
            unit = data.get("unit", "")
            if unit == "百万円":
                value_cell.number_format = '#,##0'
            elif unit == "円":
                value_cell.number_format = '#,##0'
            elif unit == "百万株":
                value_cell.number_format = '#,##0.000'
            elif unit == "年":
                value_cell.number_format = '0'  # 年度はカンマなし
            
            ws.cell(row=row, column=3, value=data["unit"]).font = unit_font
            ws.cell(row=row, column=4, value=data["desc"]).font = unit_font
            
            # 区切り線（薄いボーダー）
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = thin_border
        
        row += 1
    
    # 列幅設定
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45
    
    # 行の高さ
    for r in range(1, row):
        ws.row_dimensions[r].height = 22
    
    # 入力のヒントを追加（別シート）
    ws_help = wb.create_sheet("入力ガイド")
    help_content = [
        ["DCFバリュエーション テンプレート 入力ガイド"],
        [""],
        ["【入力方法】"],
        ["1. 「財務データ」シートの黄色セルに数値を入力してください"],
        ["2. 金額はすべて「百万円」単位で入力してください"],
        ["3. 入力後、アプリにアップロードして「データを適用」ボタンをクリック"],
        [""],
        ["【自動計算される項目】"],
        ["・売上成長率 = (売上収益 - 前期売上収益) / 前期売上収益"],
        ["・売上原価率 = 売上原価 / 売上収益"],
        ["・販管費率 = (売上収益 - 売上原価 - 営業利益) / 売上収益"],
        ["・EBITDA = 営業利益 + 減価償却費"],
        [""],
        ["【注意事項】"],
        ["・上場企業の場合はEDINET CSVを利用することで自動入力が可能です"],
        ["・非上場企業の場合は決算書から手動で入力してください"],
    ]
    for r, content in enumerate(help_content, 1):
        cell = ws_help.cell(row=r, column=1, value=content[0] if content else "")
        if r == 1:
            cell.font = Font(bold=True, size=14)
        elif content and content[0].startswith("【"):
            cell.font = Font(bold=True, size=11)
    ws_help.column_dimensions['A'].width = 70
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def read_excel_data(uploaded_file) -> tuple:
    """アップロードされたExcelから財務データを読み取る
    
    Returns:
        (HistoricalData, company_name, ratios) のタプル
        ratios: {"revenue_growth": float, "cogs_ratio": float, "sga_ratio": float}
    """
    df = pd.read_excel(uploaded_file, sheet_name="財務データ")

    # 項目名と値のマッピングを作成
    data_map = dict(zip(df["項目"], df["値"]))

    # 会社名を取得
    company_name = str(data_map.get("会社名", "")) if data_map.get("会社名") else ""
    
    # 基準年度を取得
    base_year = int(data_map.get("基準年度", 2025)) if data_map.get("基準年度") else 2025
    
    # P/L項目
    revenue = float(data_map.get("売上収益", 0))
    operating_income = float(data_map.get("営業利益", 0)) if data_map.get("営業利益") else 0
    depreciation = float(data_map.get("減価償却費", 0)) if data_map.get("減価償却費") else 0
    net_income = float(data_map.get("当期純利益", 0)) if data_map.get("当期純利益") else 0
    
    # EBITDA計算
    ebitda = operating_income + depreciation if (operating_income + depreciation) > 0 else 0

    # HistoricalDataオブジェクトを作成
    historical = HistoricalData(
        revenue=float(data_map.get("売上収益", 2080000)),
        ebitda=ebitda,
        ebit=operating_income,
        net_income=net_income,
        cash=float(data_map.get("現金及び現金同等物", 58054)),
        receivables=float(data_map.get("売上債権", 433436)),
        inventory=float(data_map.get("棚卸資産", 155938)),
        ppe=float(data_map.get("有形固定資産", 717914)),
        short_term_debt=float(data_map.get("短期借入金", 98208)),
        long_term_debt=float(data_map.get("長期借入金", 1245938)),
        shares_outstanding=float(data_map.get("発行済株式数（百万株）", 483.585)),
        current_stock_price=float(data_map.get("現在株価（円）", 4533))
    )
    
    # 比率を計算
    ratios = {
        "revenue_growth": None,
        "cogs_ratio": None,
        "sga_ratio": None,
        "base_year": base_year,
    }
    
    revenue_prior = data_map.get("前期売上収益")
    cost_of_sales = data_map.get("売上原価")
    
    # 売上成長率
    if revenue_prior and float(revenue_prior) > 0:
        ratios["revenue_growth"] = (revenue - float(revenue_prior)) / float(revenue_prior)
    
    # 売上原価率
    if cost_of_sales and revenue > 0:
        ratios["cogs_ratio"] = float(cost_of_sales) / revenue
    
    # 販管費率 = (売上収益 - 売上原価 - 営業利益) / 売上収益
    if cost_of_sales and operating_income and revenue > 0:
        sga = revenue - float(cost_of_sales) - float(operating_income)
        ratios["sga_ratio"] = sga / revenue

    return historical, company_name, ratios


def read_edinet_csv(uploaded_file) -> pd.DataFrame:
    """EDINET形式のCSVを読み込む"""
    encodings = ["utf-16", "utf-8-sig", "utf-8", "cp932", "shift_jis"]

    # 標準オプションで試行
    for encoding in encodings:
        try:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, encoding=encoding)
            return df
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception:
            continue

    # EDINET形式対応オプションで再試行（不規則なCSV対応）
    for encoding in encodings:
        try:
            uploaded_file.seek(0)
            df = pd.read_csv(
                uploaded_file,
                encoding=encoding,
                sep=None,
                engine="python",
                on_bad_lines="skip"
            )
            return df
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception:
            continue

    # すべて失敗した場合
    raise ValueError(
        "CSVファイルの文字コードを判定できませんでした。\n"
        "対応形式: UTF-16, UTF-8 (BOM付き), UTF-8, CP932, Shift-JIS"
    )


# EDINET項目名のマッピング定義（経営指標等サマリーデータ + 詳細BS項目）
# format: (項目名パターン, 相対年度フィルタ)
EDINET_ITEM_PATTERNS = {
    # P/L項目（相対年度=当期）
    "revenue": [
        ("営業収益、経営指標等", "当期"),
        ("売上高、経営指標等", "当期"),
        ("売上収益、経営指標等", "当期"),
    ],
    # 前期売上（成長率計算用）
    "revenue_prior": [
        ("営業収益、経営指標等", "前期"),
        ("売上高、経営指標等", "前期"),
        ("売上収益、経営指標等", "前期"),
    ],
    # P/L詳細項目（原価率・販管費率計算用）- 連結のみ
    "cost_of_sales": [
        ("売上原価", "当期"),
    ],
    "gross_profit": [
        ("売上総利益又は売上総損失（△）", "当期"),
        ("売上総利益", "当期"),
    ],
    "operating_income": [
        ("営業利益又は営業損失（△）", "当期"),
        ("営業利益", "当期"),
    ],
    # B/S項目（相対年度=当期末）
    "cash": [
        ("現金及び現金同等物の残高、経営指標等", "当期末"),
        ("現金及び現金同等物の期末残高", "当期末"),
    ],
    "total_assets": [
        ("総資産額、経営指標等", "当期末"),
    ],
    "net_assets": [
        ("純資産額、経営指標等", "当期末"),
    ],
    "shares_outstanding": [
        ("発行済株式総数（普通株式）、経営指標等", "当期末"),
        ("発行済株式総数、経営指標等", "当期末"),
    ],
    # 追加: 詳細BS項目（CSVに存在する場合のみ抽出）
    "receivables": [
        ("受取手形、売掛金及び契約資産", "当期末"),
        ("売掛金", "当期末"),
        ("受取手形及び売掛金", "当期末"),
    ],
    "inventory": [
        ("商品及び製品", "当期末"),
        ("棚卸資産", "当期末"),
        ("製品", "当期末"),
    ],
    "ppe": [
        ("有形固定資産", "当期末"),
    ],
    "short_term_debt": [
        ("短期借入金", "当期末"),
    ],
    "current_portion_long_term_debt": [
        ("１年内返済予定の長期借入金", "当期末"),
        ("1年内返済予定の長期借入金", "当期末"),
    ],
    "long_term_debt": [
        ("長期借入金", "当期末"),
    ],
    # 社債
    "bonds_payable": [
        ("社債", "当期末"),
    ],
    # P/L項目：支払利息（負債コスト計算用）
    "interest_expense": [
        ("支払利息", "当期"),
        ("支払利息、営業外費用", "当期"),
    ],
    # P/L項目：当期純利益（マルチプル法用）
    "net_income": [
        ("親会社株主に帰属する当期純利益又は親会社株主に帰属する当期純損失", "当期"),
        ("親会社株主に帰属する当期純利益", "当期"),
        ("当期純利益、経営指標等", "当期"),
    ],
    # P/L項目：減価償却費（EBITDA計算用）
    "depreciation": [
        ("減価償却費、販売費及び一般管理費", "当期"),
        ("減価償却費", "当期"),
    ],
    # 証券コード（株価自動取得用）
    "security_code": [
        ("証券コード、DEI", "提出日時点"),
    ],
}

# 除外する項目名（比率など数値でない項目）
EDINET_EXCLUDE_ITEMS = [
    "自己資本利益率",
    "総資産経常利益率",
    "自己資本比率",
    "増減額",
    "増減",
]

# 表示用ラベル（自動抽出項目）
EDINET_AUTO_FIELD_LABELS = {
    "revenue": "売上収益",
    "cash": "現金及び現金同等物",
    "total_assets": "総資産",
    "net_assets": "純資産",
    "shares_outstanding": "発行済株式数",
    "receivables": "売上債権",
    "inventory": "棚卸資産",
    "ppe": "有形固定資産",
    "short_term_debt": "短期借入金",
    "current_portion_long_term_debt": "1年内返済予定長期借入金",
    "long_term_debt": "長期借入金",
    "bonds_payable": "社債",
    "operating_income": "営業利益",
    "depreciation": "減価償却費",
    "net_income": "当期純利益",
}


def extract_edinet_data(df: pd.DataFrame) -> dict:
    """EDINET形式のDataFrameから財務データを抽出"""
    # EDINET CSVのカラム名を確認
    if "項目名" not in df.columns or "値" not in df.columns:
        raise ValueError("EDINET形式のCSVではありません（項目名・値カラムが必要）")

    # 抽出結果を格納
    extracted = {}

    def find_value(patterns: list, prefer_consolidated: bool = True) -> tuple:
        """項目名と相対年度で検索（連結優先）"""
        for pattern, period_filter in patterns:
            # 相対年度でフィルタ
            if "相対年度" in df.columns:
                df_filtered = df[df["相対年度"] == period_filter].copy()
            else:
                df_filtered = df.copy()

            # 完全一致を試行
            matches = df_filtered[df_filtered["項目名"] == pattern]

            # 完全一致がなければ部分一致（regex=False）
            if len(matches) == 0:
                matches = df_filtered[df_filtered["項目名"].str.contains(pattern, na=False, regex=False)]

            # 除外項目をフィルタ
            for exclude_item in EDINET_EXCLUDE_ITEMS:
                matches = matches[~matches["項目名"].str.contains(exclude_item, na=False, regex=False)]

            if len(matches) > 0:
                # 連結・個別カラムがある場合は連結を優先
                if prefer_consolidated and "連結・個別" in matches.columns:
                    consolidated = matches[matches["連結・個別"] == "連結"]
                    if len(consolidated) > 0:
                        matches = consolidated

                # 最初のマッチから数値を取得
                row = matches.iloc[0]
                val = row["値"]
                item_name = row["項目名"]
                if pd.notna(val):
                    if isinstance(val, str):
                        val_str = val.replace(",", "").replace("￥", "").replace("¥", "")
                    else:
                        val_str = str(val)
                    try:
                        val_float = float(val_str)
                        # 小数点以下が多い値（比率）は除外（1000未満で小数点あり）
                        if val_float < 1000 and "." in val_str:
                            continue
                        return val_float, item_name
                    except ValueError:
                        continue
        return None, None
    
    def find_security_code(patterns: list) -> tuple:
        """証券コードを検索（文字列として取得）"""
        for pattern, period_filter in patterns:
            if "相対年度" in df.columns:
                df_filtered = df[df["相対年度"] == period_filter].copy()
            else:
                df_filtered = df.copy()
            
            matches = df_filtered[df_filtered["項目名"].str.contains(pattern, na=False, regex=False)]
            
            if len(matches) > 0:
                row = matches.iloc[0]
                val = row["値"]
                item_name = row["項目名"]
                if pd.notna(val):
                    # 5桁の場合は先頭4桁を取得（チェックディジット除去）
                    code_str = str(val).strip()
                    if len(code_str) == 5:
                        code_str = code_str[:4]
                    return code_str, item_name
        return None, None

    # 各項目を抽出
    for field_key, patterns in EDINET_ITEM_PATTERNS.items():
        if field_key == "security_code":
            # 証券コードは文字列として取得
            value, matched_name = find_security_code(patterns)
        else:
            value, matched_name = find_value(patterns)
        extracted[field_key] = {
            "value": value,
            "matched_name": matched_name,
            "patterns": [p[0] for p in patterns]
        }

    return extracted


def calculate_financial_ratios(extracted: dict) -> dict:
    """抽出したデータから財務比率を計算"""
    ratios = {
        "revenue_growth": None,
        "cogs_ratio": None,
        "sga_ratio": None,
    }
    
    # 売上成長率の計算
    revenue = extracted.get("revenue", {}).get("value")
    revenue_prior = extracted.get("revenue_prior", {}).get("value")
    if revenue and revenue_prior and revenue_prior > 0:
        ratios["revenue_growth"] = (revenue - revenue_prior) / revenue_prior
    
    # 売上原価率・販管費率の計算
    cost_of_sales = extracted.get("cost_of_sales", {}).get("value")
    gross_profit = extracted.get("gross_profit", {}).get("value")
    operating_income = extracted.get("operating_income", {}).get("value")
    
    # 売上高 = 売上原価 + 売上総利益
    if cost_of_sales and gross_profit:
        sales = cost_of_sales + gross_profit
        if sales > 0:
            # 売上原価率
            ratios["cogs_ratio"] = cost_of_sales / sales
            
            # 販管費率 = (売上総利益 - 営業利益) / 売上高
            if operating_income is not None:
                sga = gross_profit - operating_income
                ratios["sga_ratio"] = sga / sales
    
    return ratios


def create_historical_from_edinet(
    extracted: dict,
    manual_inputs: dict,
    current_stock_price: float = 4533
) -> HistoricalData:
    """抽出したEDINETデータと手動入力からHistoricalDataを作成
    
    注意: EDINET CSVの値は「円」単位なので、百万円に変換する
    """
    def get_val(key: str, default: float, convert_to_millions: bool = True) -> float:
        # まず手動入力をチェック（0以外の値がある場合）- 既に百万円単位
        if key in manual_inputs and manual_inputs[key] is not None and manual_inputs[key] > 0:
            return manual_inputs[key]
        # 次に抽出データをチェック - 円単位なので百万円に変換
        if extracted.get(key) and extracted[key]["value"] is not None:
            val = extracted[key]["value"]
            if convert_to_millions:
                return val / 1_000_000  # 円 → 百万円
            else:
                return val
        return default
    
    # 有利子負債の合算（1年内返済分も含める）
    short_term = get_val("short_term_debt", 0)
    current_portion = get_val("current_portion_long_term_debt", 0)
    long_term = get_val("long_term_debt", 0)
    bonds = get_val("bonds_payable", 0)
    
    # 短期有利子負債 = 短期借入金 + 1年内返済予定の長期借入金
    total_short_term_debt = short_term + current_portion if (short_term + current_portion) > 0 else 98208
    # 長期有利子負債 = 長期借入金 + 社債
    total_long_term_debt = long_term + bonds if (long_term + bonds) > 0 else 1245938
    
    # P/L項目（マルチプル法用）
    operating_income = get_val("operating_income", 0)
    depreciation = get_val("depreciation", 0)
    net_income = get_val("net_income", 0)
    
    # EBITDA = 営業利益 + 減価償却費
    ebitda = operating_income + depreciation if (operating_income + depreciation) > 0 else 0
    # EBIT = 営業利益
    ebit = operating_income if operating_income > 0 else 0

    return HistoricalData(
        revenue=get_val("revenue", 2080000),
        ebitda=ebitda,
        ebit=ebit,
        net_income=net_income,
        cash=get_val("cash", 58054),
        receivables=get_val("receivables", 433436),
        inventory=get_val("inventory", 155938),
        ppe=get_val("ppe", 717914),
        short_term_debt=total_short_term_debt,
        long_term_debt=total_long_term_debt,
        shares_outstanding=get_val("shares_outstanding", 483.585),  # 株 → 百万株
        current_stock_price=current_stock_price
    )


def create_analysis_excel(model: DCFModel, inputs: dict) -> BytesIO:
    """分析結果をExcelファイルとして作成"""
    output = BytesIO()

    # スタイル定義
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    sub_header_fill = PatternFill(start_color="C9A962", end_color="C9A962", fill_type="solid")
    sub_header_font = Font(bold=True, size=11)
    number_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # === シート1: サマリー ===
        summary = model.summary()
        perp = summary["valuation"]["perpetuity"]
        exit_m = summary["valuation"]["exit_multiple"]

        summary_data = {
            "項目": [
                "【バリュエーション結果】",
                "企業価値（永久成長率法）",
                "企業価値（Exit Multiple法）",
                "株式価値（永久成長率法）",
                "株式価値（Exit Multiple法）",
                "理論株価（永久成長率法）",
                "理論株価（Exit Multiple法）",
                "現在株価",
                "アップサイド（永久成長率法）",
                "",
                "【WACC構成要素】",
                "リスクフリーレート",
                "アンレバードβ",
                "レバードβ",
                "株式リスクプレミアム",
                "サイズプレミアム",
                "株主資本コスト",
                "負債コスト（税前）",
                "負債コスト（税後）",
                "目標D/Eレシオ",
                "WACC",
                "",
                "【主要前提条件】",
                "シナリオ",
                "売上成長率",
                "売上原価率",
                "販管費率",
                "永久成長率",
                "Exit Multiple",
            ],
            "値": [
                "",
                f"{perp['enterprise_value']:,.0f}",
                f"{exit_m['enterprise_value']:,.0f}",
                f"{perp['equity_value']:,.0f}",
                f"{exit_m['equity_value']:,.0f}",
                f"{perp['price_per_share']:,.0f}",
                f"{exit_m['price_per_share']:,.0f}",
                f"{model.historical.current_stock_price:,.0f}",
                f"{(perp['price_per_share'] / model.historical.current_stock_price - 1) * 100:.1f}%",
                "",
                "",
                f"{model.assumptions.risk_free_rate:.2%}",
                f"{model.assumptions.unlevered_beta:.2f}",
                f"{model.levered_beta:.2f}",
                f"{model.assumptions.equity_risk_premium:.2%}",
                f"{model.assumptions.size_premium:.2%}",
                f"{model.cost_of_equity:.2%}",
                f"{model.assumptions.cost_of_debt:.2%}",
                f"{model.assumptions.cost_of_debt * (1 - 0.30):.2%}",  # 実効税率30%
                f"{model.assumptions.target_de_ratio:.2f}",
                f"{model.wacc:.2%}",
                "",
                "",
                inputs["scenario"].value,
                f"{inputs['revenue_growth']:.1%}",
                f"{inputs['cogs_ratio']:.1%}",
                f"{inputs['sga_ratio']:.1%}",
                f"{inputs['terminal_growth_rate']:.2%}",
                f"{inputs['exit_ebitda_multiple']:.1f}x",
            ],
            "単位": [
                "",
                "百万円",
                "百万円",
                "百万円",
                "百万円",
                "円",
                "円",
                "円",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        }

        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, index=False, sheet_name="サマリー")

        # サマリーシートのスタイル適用
        ws_summary = writer.sheets["サマリー"]
        ws_summary.column_dimensions["A"].width = 35
        ws_summary.column_dimensions["B"].width = 20
        ws_summary.column_dimensions["C"].width = 10

        # ヘッダー行のスタイル
        for col in range(1, 4):
            cell = ws_summary.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # セクションヘッダー行のスタイル
        section_rows = [2, 12, 24]  # 【】で始まる行
        for row in range(2, len(df_summary) + 2):
            for col in range(1, 4):
                cell = ws_summary.cell(row=row, column=col)
                cell.border = thin_border
                if row - 1 in [0, 10, 22]:  # セクションヘッダー
                    cell.fill = sub_header_fill
                    cell.font = sub_header_font
                elif col == 2:
                    cell.font = number_font
                    cell.alignment = Alignment(horizontal="right")

        # === シート2: FCF予測 ===
        df_fcf = model.projection_table()
        df_fcf_export = df_fcf.reset_index()
        df_fcf_export.columns = ["項目"] + list(df_fcf.columns)
        df_fcf_export.to_excel(writer, index=False, sheet_name="FCF予測")

        ws_fcf = writer.sheets["FCF予測"]
        ws_fcf.column_dimensions["A"].width = 15

        # ヘッダー行のスタイル
        for col in range(1, len(df_fcf_export.columns) + 1):
            cell = ws_fcf.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if col > 1:
                ws_fcf.column_dimensions[cell.column_letter].width = 14

        # データセルのスタイル
        for row in range(2, len(df_fcf_export) + 2):
            for col in range(1, len(df_fcf_export.columns) + 1):
                cell = ws_fcf.cell(row=row, column=col)
                cell.border = thin_border
                if col > 1:
                    cell.font = number_font
                    cell.alignment = Alignment(horizontal="right")
                    # 数値フォーマット
                    if isinstance(cell.value, (int, float)):
                        if df_fcf_export.iloc[row - 2, 0] == "割引係数":
                            cell.number_format = "0.0000"
                        else:
                            cell.number_format = "#,##0"

        # === シート3: 感応度分析_永久成長率法 ===
        df_sens_perp = model.sensitivity_analysis("perpetuity")
        df_sens_perp_export = df_sens_perp.reset_index()
        df_sens_perp_export.columns = ["永久成長率\\WACC"] + list(df_sens_perp.columns)
        df_sens_perp_export.to_excel(writer, index=False, sheet_name="感応度分析_永久成長率法")

        ws_sens_perp = writer.sheets["感応度分析_永久成長率法"]
        ws_sens_perp.column_dimensions["A"].width = 18

        # ヘッダー行のスタイル
        for col in range(1, len(df_sens_perp_export.columns) + 1):
            cell = ws_sens_perp.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if col > 1:
                ws_sens_perp.column_dimensions[cell.column_letter].width = 12

        # データセルのスタイル
        for row in range(2, len(df_sens_perp_export) + 2):
            for col in range(1, len(df_sens_perp_export.columns) + 1):
                cell = ws_sens_perp.cell(row=row, column=col)
                cell.border = thin_border
                if col == 1:
                    cell.fill = sub_header_fill
                    cell.font = sub_header_font
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.font = number_font
                    cell.alignment = Alignment(horizontal="right")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"

        # === シート4: 感応度分析_Exit Multiple法 ===
        df_sens_exit = model.sensitivity_analysis("exit_multiple")
        df_sens_exit_export = df_sens_exit.reset_index()
        df_sens_exit_export.columns = ["Exit Multiple\\WACC"] + list(df_sens_exit.columns)
        df_sens_exit_export.to_excel(writer, index=False, sheet_name="感応度分析_Exit Multiple法")

        ws_sens_exit = writer.sheets["感応度分析_Exit Multiple法"]
        ws_sens_exit.column_dimensions["A"].width = 18

        # ヘッダー行のスタイル
        for col in range(1, len(df_sens_exit_export.columns) + 1):
            cell = ws_sens_exit.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if col > 1:
                ws_sens_exit.column_dimensions[cell.column_letter].width = 12

        # データセルのスタイル
        for row in range(2, len(df_sens_exit_export) + 2):
            for col in range(1, len(df_sens_exit_export.columns) + 1):
                cell = ws_sens_exit.cell(row=row, column=col)
                cell.border = thin_border
                if col == 1:
                    cell.fill = sub_header_fill
                    cell.font = sub_header_font
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.font = number_font
                    cell.alignment = Alignment(horizontal="right")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"

    output.seek(0)
    return output


def register_japanese_font():
    """日本語フォントを登録（システムにあるフォントを探す）"""
    # フォント検索パス（優先順）
    font_paths = [
        # macOS
        "/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc",
        "/System/Library/Fonts/Hiragino Sans GB.ttc",
        "/Library/Fonts/Arial Unicode.ttf",
        # Linux
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/google-noto-cjk/NotoSansCJK-Regular.ttc",
        # Windows
        "C:/Windows/Fonts/msgothic.ttc",
        "C:/Windows/Fonts/meiryo.ttc",
        "C:/Windows/Fonts/YuGothM.ttc",
    ]
    
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('JapaneseFont', font_path))
                return 'JapaneseFont'
            except Exception:
                continue
    
    # フォントが見つからない場合はデフォルトフォントを使用
    return 'Helvetica'


def create_analysis_pdf(model: DCFModel, inputs: dict) -> BytesIO:
    """分析結果をPDFファイルとして作成"""
    output = BytesIO()
    
    # 日本語フォント登録
    font_name = register_japanese_font()
    
    # PDFドキュメント作成
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )
    
    # スタイル定義
    styles = getSampleStyleSheet()
    
    # カスタムスタイル
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontName=font_name,
        fontSize=18,
        spaceAfter=20,
        textColor=colors.HexColor('#1E3A5F')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontName=font_name,
        fontSize=12,
        spaceBefore=15,
        spaceAfter=10,
        textColor=colors.HexColor('#1E3A5F')
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=9
    )
    
    # コンテンツ
    story = []
    
    # タイトル
    company_name = st.session_state.get("company_name", "")
    title_text = f"{company_name} DCF Valuation Report" if company_name else "DCF Valuation Report"
    story.append(Paragraph(title_text, title_style))
    story.append(Paragraph(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d')}", normal_style))
    story.append(Spacer(1, 10*mm))
    
    # サマリーデータ取得
    summary = model.summary()
    perp = summary["valuation"]["perpetuity"]
    exit_m = summary["valuation"]["exit_multiple"]
    
    # === セクション1: バリュエーション結果 ===
    story.append(Paragraph("Valuation Summary", heading_style))
    
    valuation_data = [
        ["Item", "Perpetuity Method", "Exit Multiple Method"],
        ["Enterprise Value (MM JPY)", f"{perp['enterprise_value']:,.0f}", f"{exit_m['enterprise_value']:,.0f}"],
        ["Equity Value (MM JPY)", f"{perp['equity_value']:,.0f}", f"{exit_m['equity_value']:,.0f}"],
        ["Price per Share (JPY)", f"{perp['price_per_share']:,.0f}", f"{exit_m['price_per_share']:,.0f}"],
        ["Current Price (JPY)", f"{model.historical.current_stock_price:,.0f}", f"{model.historical.current_stock_price:,.0f}"],
        ["Upside", f"{(perp['price_per_share'] / model.historical.current_stock_price - 1) * 100:+.1f}%", 
                   f"{(exit_m['price_per_share'] / model.historical.current_stock_price - 1) * 100:+.1f}%"],
    ]
    
    val_table = Table(valuation_data, colWidths=[70*mm, 45*mm, 45*mm])
    val_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#E8EEF4')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(val_table)
    story.append(Spacer(1, 8*mm))
    
    # === セクション2: WACC構成要素 ===
    story.append(Paragraph("WACC Components", heading_style))
    
    wacc_data = [
        ["Component", "Value"],
        ["Risk-free Rate", f"{model.assumptions.risk_free_rate:.2%}"],
        ["Unlevered Beta", f"{model.assumptions.unlevered_beta:.2f}"],
        ["Levered Beta", f"{model.levered_beta:.2f}"],
        ["Equity Risk Premium", f"{model.assumptions.equity_risk_premium:.2%}"],
        ["Size Premium", f"{model.assumptions.size_premium:.2%}"],
        ["Cost of Equity", f"{model.cost_of_equity:.2%}"],
        ["Cost of Debt (Pre-tax)", f"{model.assumptions.cost_of_debt:.2%}"],
        ["Cost of Debt (After-tax)", f"{model.assumptions.cost_of_debt * 0.70:.2%}"],
        ["Target D/E Ratio", f"{model.assumptions.target_de_ratio:.2f}"],
        ["WACC", f"{model.wacc:.2%}"],
    ]
    
    wacc_table = Table(wacc_data, colWidths=[80*mm, 40*mm])
    wacc_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#E8EEF4')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(wacc_table)
    story.append(Spacer(1, 8*mm))
    
    # === セクション3: 主要前提条件 ===
    story.append(Paragraph("Key Assumptions", heading_style))
    
    # シナリオ名の変換
    scenario_name_map = {
        "強気": "Bull",
        "ベース": "Base",
        "弱気": "Bear",
    }
    scenario_name = scenario_name_map.get(inputs["scenario"].value, inputs["scenario"].value)
    
    assumptions_data = [
        ["Item", "Value"],
        ["Scenario", scenario_name],
        ["Revenue Growth Rate", f"{inputs['revenue_growth']:.1%}"],
        ["COGS Ratio", f"{inputs['cogs_ratio']:.1%}"],
        ["SG&A Ratio", f"{inputs['sga_ratio']:.1%}"],
        ["Terminal Growth Rate", f"{inputs['terminal_growth_rate']:.2%}"],
        ["Exit EBITDA Multiple", f"{inputs['exit_ebitda_multiple']:.1f}x"],
    ]
    
    assump_table = Table(assumptions_data, colWidths=[80*mm, 40*mm])
    assump_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#E8EEF4')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(assump_table)
    
    # 改ページ
    story.append(PageBreak())
    
    # === セクション4: FCF予測 ===
    story.append(Paragraph("FCF Projection (MM JPY)", heading_style))
    
    df_projection = model.projection_table()
    
    # 日本語→英語の項目名マッピング
    row_name_map = {
        "売上収益": "Revenue",
        "EBITDA": "EBITDA",
        "減価償却": "D&A",
        "EBIT": "EBIT",
        "NOPAT": "NOPAT",
        "設備投資": "CapEx",
        "Δ運転資本": "Chg in NWC",
        "FCF": "FCF",
        "割引係数": "Discount Factor",
        "FCF現在価値": "PV of FCF",
    }
    
    # FCF予測テーブル
    fcf_headers = ["Item"] + list(df_projection.columns)
    fcf_data = [fcf_headers]
    
    for idx in df_projection.index:
        row_name = row_name_map.get(idx, idx)
        row = [row_name]
        for col in df_projection.columns:
            val = df_projection.loc[idx, col]
            if idx == "割引係数":
                row.append(f"{val:.4f}" if pd.notna(val) else "-")
            else:
                row.append(f"{val:,.0f}" if pd.notna(val) else "-")
        fcf_data.append(row)
    
    # カラム幅（項目名を広く、年度は均等に）
    col_widths = [35*mm] + [25*mm] * len(df_projection.columns)
    
    fcf_table = Table(fcf_data, colWidths=col_widths)
    fcf_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#E8EEF4')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(fcf_table)
    story.append(Spacer(1, 8*mm))
    
    # === セクション5: 感応度分析 ===
    story.append(Paragraph("Sensitivity Analysis - Perpetuity Method (JPY)", heading_style))
    
    df_sens = model.sensitivity_analysis("perpetuity")
    
    # 感応度テーブル
    sens_headers = ["Growth\\WACC"] + list(df_sens.columns)
    sens_data = [sens_headers]
    
    for idx in df_sens.index:
        row = [idx]
        for col in df_sens.columns:
            val = df_sens.loc[idx, col]
            row.append(f"{val:,.0f}" if pd.notna(val) else "-")
        sens_data.append(row)
    
    sens_col_widths = [25*mm] + [22*mm] * len(df_sens.columns)
    
    sens_table = Table(sens_data, colWidths=sens_col_widths)
    sens_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#C9A962')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(sens_table)
    
    # フッター情報
    story.append(Spacer(1, 15*mm))
    footer_text = f"Generated by DCF Valuation Model | {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    story.append(Paragraph(footer_text, ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=8,
        textColor=colors.grey,
        alignment=1  # Center
    )))
    
    # PDF生成
    doc.build(story)
    output.seek(0)
    return output


def create_model_from_inputs(
    scenario: Scenario,
    base_year: int,
    revenue_growth: float,
    cogs_ratio: float,
    sga_ratio: float,
    risk_free_rate: float,
    unlevered_beta: float,
    target_de_ratio: float,
    equity_risk_premium: float,
    size_premium: float,
    cost_of_debt: float,
    terminal_growth_rate: float,
    exit_ebitda_multiple: float,
    historical_data: HistoricalData = None
) -> DCFModel:
    """入力値からモデルを作成"""
    assumptions = Assumptions(
        scenario=scenario,
        base_year=base_year,
        revenue_growth={
            Scenario.BULL: revenue_growth if scenario == Scenario.BULL else 0.15,
            Scenario.BASE: revenue_growth if scenario == Scenario.BASE else 0.10,
            Scenario.BEAR: revenue_growth if scenario == Scenario.BEAR else 0.05
        },
        cogs_ratio=cogs_ratio,
        sga_ratio=sga_ratio,
        risk_free_rate=risk_free_rate,
        unlevered_beta=unlevered_beta,
        target_de_ratio=target_de_ratio,
        equity_risk_premium=equity_risk_premium,
        size_premium=size_premium,
        cost_of_debt=cost_of_debt,
        terminal_growth_rate=terminal_growth_rate,
        exit_ebitda_multiple=exit_ebitda_multiple
    )
    return DCFModel(assumptions=assumptions, historical=historical_data)


def render_header():
    """ヘッダー表示"""
    # セッションステートから会社名を取得（未設定時は空）
    company_name = st.session_state.get("company_name", "")
    
    # 会社名がある場合は「会社名 DCFバリュエーション」、なければ「DCFバリュエーション」
    title = f"{company_name} DCFバリュエーション" if company_name else "DCFバリュエーション"
    
    st.markdown(f"""
    <div class="main-header">
        <h1>{title}</h1>
        <p>Discounted Cash Flow Valuation Model</p>
    </div>
    """, unsafe_allow_html=True)


def reset_session_state():
    """セッションステートをリセット（新しい企業分析用）"""
    keys_to_reset = [
        "company_name",
        "calculated_revenue_growth",
        "calculated_cogs_ratio", 
        "calculated_sga_ratio",
        "calculated_cost_of_debt",
        "cost_of_debt_status",
        "base_year",
        "fetched_stock_price",
        "fetched_company_name",
        "edinet_extracted",
        "edinet_manual_inputs",
        "csv_historical_data",
        "current_excel_file",
        "current_csv_file",
    ]
    for key in keys_to_reset:
        if key in st.session_state:
            del st.session_state[key]


def render_sidebar() -> dict:
    """サイドバー（入力パネル）"""
    with st.sidebar:
        # === Excelアップロード（最上部） ===
        st.markdown('<div class="sidebar-header">財務データ入力</div>', unsafe_allow_html=True)
        
        # リセットボタン
        if st.button("🔄 データをリセット", help="現在のデータをクリアして新しい企業を分析"):
            reset_session_state()
            st.rerun()

        # テンプレートダウンロード
        template_file = create_excel_template()
        st.download_button(
            label="テンプレートをダウンロード",
            data=template_file,
            file_name="dcf_input_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Excelテンプレートをダウンロードして、財務データを入力してください"
        )

        # Excelアップロード
        uploaded_file = st.file_uploader(
            "Excelファイルをアップロード",
            type=["xlsx", "xls"],
            help="テンプレートに財務データを入力してアップロードしてください",
            key="excel_uploader"
        )

        # アップロードされたファイルからデータを読み込む
        historical_data = None
        if uploaded_file is not None:
            # ファイル名が変わったら自動リセット
            current_file_name = uploaded_file.name
            previous_file_name = st.session_state.get("current_excel_file")
            
            if previous_file_name and previous_file_name != current_file_name:
                # 新しいファイルがアップロードされた
                reset_session_state()
                st.info(f"新しいファイル「{current_file_name}」を読み込みます")
            
            st.session_state["current_excel_file"] = current_file_name
            
            try:
                historical_data, excel_company_name, excel_ratios = read_excel_data(uploaded_file)
                
                # 会社名をセッションステートに保存
                if excel_company_name:
                    st.session_state["company_name"] = excel_company_name
                
                # 計算された比率をセッションステートに保存
                if excel_ratios["revenue_growth"] is not None:
                    st.session_state["calculated_revenue_growth"] = excel_ratios["revenue_growth"]
                if excel_ratios["cogs_ratio"] is not None:
                    st.session_state["calculated_cogs_ratio"] = excel_ratios["cogs_ratio"]
                if excel_ratios["sga_ratio"] is not None:
                    st.session_state["calculated_sga_ratio"] = excel_ratios["sga_ratio"]
                
                # 基準年度をセッションステートに保存
                if excel_ratios.get("base_year"):
                    st.session_state["base_year"] = excel_ratios["base_year"]
                
                st.success("財務データを読み込みました")

                # 読み込んだデータをプレビュー表示
                with st.expander("読み込んだデータ", expanded=False):
                    preview_items = ["会社名", "売上収益", "EBITDA", "当期純利益", "現金", "売上債権", 
                                    "棚卸資産", "有形固定資産", "短期借入金", "長期借入金", 
                                    "発行済株式数", "現在株価"]
                    preview_values = [
                        excel_company_name if excel_company_name else "（未設定）",
                        f"{historical_data.revenue:,.0f}百万円",
                        f"{historical_data.ebitda:,.0f}百万円" if historical_data.ebitda > 0 else "（未入力）",
                        f"{historical_data.net_income:,.0f}百万円" if historical_data.net_income > 0 else "（未入力）",
                        f"{historical_data.cash:,.0f}百万円",
                        f"{historical_data.receivables:,.0f}百万円",
                        f"{historical_data.inventory:,.0f}百万円",
                        f"{historical_data.ppe:,.0f}百万円",
                        f"{historical_data.short_term_debt:,.0f}百万円",
                        f"{historical_data.long_term_debt:,.0f}百万円",
                        f"{historical_data.shares_outstanding:.3f}百万株",
                        f"{historical_data.current_stock_price:,.0f}円"
                    ]
                    preview_data = {"項目": preview_items, "値": preview_values}
                    st.dataframe(pd.DataFrame(preview_data), hide_index=True)
                    
                    # 計算された比率を表示
                    if any(v is not None for v in excel_ratios.values()):
                        st.markdown("**自動計算された比率**")
                        ratio_rows = []
                        if excel_ratios["revenue_growth"] is not None:
                            ratio_rows.append({"項目": "売上成長率", "値": f"{excel_ratios['revenue_growth']:.1%}"})
                        if excel_ratios["cogs_ratio"] is not None:
                            ratio_rows.append({"項目": "売上原価率", "値": f"{excel_ratios['cogs_ratio']:.1%}"})
                        if excel_ratios["sga_ratio"] is not None:
                            ratio_rows.append({"項目": "販管費率", "値": f"{excel_ratios['sga_ratio']:.1%}"})
                        st.dataframe(pd.DataFrame(ratio_rows), hide_index=True)
            except Exception as e:
                st.error(f"ファイルの読み込みに失敗しました: {str(e)}")

        # === EDINET CSV読み込み ===
        st.markdown('<div class="sidebar-header">EDINET CSV読み込み</div>', unsafe_allow_html=True)

        csv_file = st.file_uploader(
            "CSVファイルをアップロード",
            type=["csv"],
            help="EDINET形式のCSVファイル（UTF-16推奨）",
            key="csv_uploader"
        )

        csv_historical_data = None
        if csv_file is not None:
            # ファイル名が変わったら自動リセット
            current_csv_name = csv_file.name
            previous_csv_name = st.session_state.get("current_csv_file")
            
            if previous_csv_name and previous_csv_name != current_csv_name:
                # 新しいCSVファイルがアップロードされた
                reset_session_state()
                st.info(f"新しいファイル「{current_csv_name}」を読み込みます")
            
            st.session_state["current_csv_file"] = current_csv_name
            
            try:
                # CSVを読み込み（エンコーディング自動判定）
                df_csv = read_edinet_csv(csv_file)

                st.success(f"CSVを読み込みました（{len(df_csv)}行 × {len(df_csv.columns)}列）")

                # 生データプレビュー
                with st.expander("生データプレビュー", expanded=False):
                    st.dataframe(df_csv.head(10), use_container_width=True)

                # EDINET形式かどうかを判定
                is_edinet = "項目名" in df_csv.columns and "値" in df_csv.columns

                if is_edinet:
                    # EDINET形式：自動抽出
                    st.markdown("**自動抽出結果**")

                    extracted = extract_edinet_data(df_csv)
                    
                    # 財務比率を計算
                    ratios = calculate_financial_ratios(extracted)
                    
                    # 計算結果をセッションステートに保存
                    if ratios["revenue_growth"] is not None:
                        st.session_state["calculated_revenue_growth"] = ratios["revenue_growth"]
                    if ratios["cogs_ratio"] is not None:
                        st.session_state["calculated_cogs_ratio"] = ratios["cogs_ratio"]
                    if ratios["sga_ratio"] is not None:
                        st.session_state["calculated_sga_ratio"] = ratios["sga_ratio"]
                    
                    # 負債コストを計算
                    cod, cod_status = calculate_cost_of_debt(extracted)
                    if cod is not None:
                        st.session_state["calculated_cost_of_debt"] = cod
                        st.session_state["cost_of_debt_status"] = cod_status

                    # 抽出結果をテーブル表示（百万円に変換）
                    preview_rows = []
                    for key, label in EDINET_AUTO_FIELD_LABELS.items():
                        data = extracted.get(key, {})
                        value = data.get("value")
                        matched = data.get("matched_name", "—")
                        if value is not None:
                            # 円 → 百万円に変換して表示
                            value_in_millions = value / 1_000_000
                            if key == "shares_outstanding":
                                preview_rows.append({
                                    "項目": label,
                                    "マッチした項目名": matched,
                                    "値（百万円/百万株）": f"{value_in_millions:,.3f}"
                                })
                            else:
                                preview_rows.append({
                                    "項目": label,
                                    "マッチした項目名": matched,
                                    "値（百万円/百万株）": f"{value_in_millions:,.0f}"
                                })
                        else:
                            preview_rows.append({
                                "項目": label,
                                "マッチした項目名": "（未検出）",
                                "値（百万円/百万株）": "—"
                            })

                    st.dataframe(pd.DataFrame(preview_rows), hide_index=True, use_container_width=True)
                    
                    # 計算された財務比率を表示
                    if any(v is not None for v in ratios.values()):
                        st.markdown("**自動計算された比率**")
                        ratio_rows = []
                        if ratios["revenue_growth"] is not None:
                            ratio_rows.append({
                                "項目": "売上成長率（前期比）",
                                "値": f"{ratios['revenue_growth']:.1%}"
                            })
                        if ratios["cogs_ratio"] is not None:
                            ratio_rows.append({
                                "項目": "売上原価率",
                                "値": f"{ratios['cogs_ratio']:.1%}"
                            })
                        if ratios["sga_ratio"] is not None:
                            ratio_rows.append({
                                "項目": "販管費率",
                                "値": f"{ratios['sga_ratio']:.1%}"
                            })
                        st.dataframe(pd.DataFrame(ratio_rows), hide_index=True, use_container_width=True)
                        st.caption("※ これらの比率は収益予測の初期値として自動設定されます")

                    # 未検出項目のみ手動入力フィールドを表示
                    missing_fields = {
                        k: v for k, v in EDINET_AUTO_FIELD_LABELS.items()
                        if extracted.get(k, {}).get("value") is None
                        and k not in ["revenue", "total_assets", "net_assets"]  # 必須項目以外
                    }

                    manual_inputs = {}
                    if missing_fields:
                        st.markdown("**手動入力（百万円）**")
                        st.caption("以下の項目はCSVから検出できなかったため手動入力してください")

                        for field_key, label in missing_fields.items():
                            manual_inputs[field_key] = st.number_input(
                                label,
                                min_value=0.0,
                                value=0.0,
                                step=1000.0,
                                format="%.0f",
                                key=f"manual_{field_key}"
                            )

                    # 現在株価入力（自動取得対応）
                    st.markdown("**株価情報**")
                    
                    # CSVから証券コードを取得
                    csv_security_code = extracted.get("security_code", {}).get("value")
                    
                    # 証券コード入力（CSV取得値をデフォルトに）
                    col_ticker, col_btn = st.columns([2, 1])
                    with col_ticker:
                        ticker_code = st.text_input(
                            "証券コード（4桁）",
                            value=csv_security_code if csv_security_code else "",
                            placeholder="例: 8273",
                            help="東証の証券コード4桁（CSVから自動取得）"
                        )
                    
                    with col_btn:
                        st.write("")  # スペーサー
                        fetch_clicked = st.button("株価取得", use_container_width=True)
                    
                    # 自動株価取得（CSVから証券コードが取得でき、まだ株価を取得していない場合）
                    auto_fetch = False
                    if csv_security_code and "fetched_stock_price" not in st.session_state:
                        auto_fetch = True
                    
                    # 株価取得処理
                    if (fetch_clicked or auto_fetch) and ticker_code:
                        with st.spinner("株価を取得中..."):
                            price, result = get_stock_price(ticker_code)
                            if price is not None:
                                st.session_state["fetched_stock_price"] = price
                                st.session_state["fetched_company_name"] = result
                                # 会社名をヘッダー用にも保存
                                st.session_state["company_name"] = result
                                if auto_fetch:
                                    st.success(f"株価を自動取得: {result} - {price:,.0f}円")
                                else:
                                    st.success(f"取得成功: {result} - {price:,.0f}円")
                            else:
                                st.warning(f"株価取得失敗: {result}（手動入力してください）")
                    
                    # 取得した株価があれば表示、なければ手動入力
                    default_price = st.session_state.get("fetched_stock_price", 0.0)
                    current_price = st.number_input(
                        "現在株価（円）",
                        min_value=0.0,
                        value=float(default_price),
                        step=1.0,
                        help="証券コードを入力して「株価取得」を押すか、手動で入力してください"
                    )

                    # セッションに抽出データを保存
                    st.session_state["edinet_extracted"] = extracted
                    st.session_state["edinet_manual_inputs"] = manual_inputs

                    # データを適用ボタン
                    if st.button("データを適用", type="primary", use_container_width=True, key="apply_edinet"):
                        csv_historical_data = create_historical_from_edinet(extracted, manual_inputs, current_price)
                        st.session_state["csv_historical_data"] = csv_historical_data
                        st.success("EDINETデータを適用しました")
                        st.rerun()  # 会社名を反映するためにリロード

                else:
                    # 非EDINET形式：カラム一覧を表示
                    st.warning("EDINET形式ではありません（項目名・値カラムが必要）")
                    st.markdown(f"検出されたカラム: {', '.join(df_csv.columns)}")

                # セッションステートからデータを取得
                if "csv_historical_data" in st.session_state:
                    csv_historical_data = st.session_state["csv_historical_data"]

                    # 適用済みデータのプレビュー
                    with st.expander("適用済みデータ", expanded=False):
                        # 有利子負債合計とネットデット計算
                        total_debt = csv_historical_data.short_term_debt + csv_historical_data.long_term_debt
                        net_debt = total_debt - csv_historical_data.cash
                        
                        applied_data = {
                            "項目": ["売上収益", "現金", "売上債権", "棚卸資産", "有形固定資産",
                                    "短期有利子負債", "長期有利子負債", "【有利子負債合計】", "【ネットデット】",
                                    "発行済株式数", "現在株価"],
                            "値": [
                                f"{csv_historical_data.revenue:,.0f}百万円",
                                f"{csv_historical_data.cash:,.0f}百万円",
                                f"{csv_historical_data.receivables:,.0f}百万円",
                                f"{csv_historical_data.inventory:,.0f}百万円",
                                f"{csv_historical_data.ppe:,.0f}百万円",
                                f"{csv_historical_data.short_term_debt:,.0f}百万円",
                                f"{csv_historical_data.long_term_debt:,.0f}百万円",
                                f"{total_debt:,.0f}百万円",
                                f"{net_debt:,.0f}百万円",
                                f"{csv_historical_data.shares_outstanding:.3f}百万株",
                                f"{csv_historical_data.current_stock_price:,.0f}円"
                            ]
                        }
                        st.dataframe(pd.DataFrame(applied_data), hide_index=True)

            except Exception as e:
                st.error(f"CSVの読み込みに失敗しました: {str(e)}")

        # CSVデータがあればExcelデータより優先（セッションステートから取得）
        if csv_historical_data is not None:
            historical_data = csv_historical_data
        elif "csv_historical_data" in st.session_state:
            # st.rerun()後でもセッションステートから取得
            historical_data = st.session_state["csv_historical_data"]

        st.markdown("---")

        # === シナリオ選択 ===
        st.markdown('<div class="sidebar-header">シナリオ選択</div>', unsafe_allow_html=True)
        scenario_str = st.radio(
            "シナリオ",
            options=["ベース", "強気", "弱気"],
            horizontal=True,
            label_visibility="collapsed"
        )
        scenario_map = {"強気": Scenario.BULL, "ベース": Scenario.BASE, "弱気": Scenario.BEAR}
        scenario = scenario_map[scenario_str]

        # 基準年度（CSVから自動取得または手動入力）
        base_year = st.number_input(
            "基準年度",
            min_value=2015,
            max_value=2030,
            value=st.session_state.get("base_year", 2025),
            step=1,
            help="財務データの基準年度（当期）を入力してください"
        )

        # デフォルト成長率（シナリオ別）
        default_growth = {Scenario.BULL: 15.0, Scenario.BASE: 10.0, Scenario.BEAR: 5.0}
        
        # EDINET CSVから計算された値があれば使用、なければデフォルト
        calculated_growth = st.session_state.get("calculated_revenue_growth")
        if calculated_growth is not None:
            # 計算値をパーセント表示用に変換（0-30の範囲にクリップ）
            growth_default = min(max(calculated_growth * 100, 0.0), 30.0)
        else:
            growth_default = default_growth[scenario]
        
        calculated_cogs = st.session_state.get("calculated_cogs_ratio")
        cogs_default = calculated_cogs * 100 if calculated_cogs is not None else 62.5
        cogs_default = min(max(cogs_default, 50.0), 80.0)  # 範囲内にクリップ
        
        calculated_sga = st.session_state.get("calculated_sga_ratio")
        sga_default = calculated_sga * 100 if calculated_sga is not None else 28.0
        sga_default = min(max(sga_default, 15.0), 40.0)  # 範囲内にクリップ

        st.markdown('<div class="sidebar-header">収益予測</div>', unsafe_allow_html=True)
        
        # 自動計算値がある場合は注記を表示
        if calculated_growth is not None or calculated_cogs is not None:
            st.caption("※ EDINETデータから自動計算された初期値を使用中")
        
        revenue_growth = st.slider(
            "売上成長率 (%)",
            min_value=0.0, max_value=30.0,
            value=growth_default,
            step=0.5,
            format="%.1f%%"
        ) / 100

        cogs_ratio = st.slider(
            "売上原価率 (%)",
            min_value=50.0, max_value=80.0,
            value=cogs_default,
            step=0.5,
            format="%.1f%%"
        ) / 100

        sga_ratio = st.slider(
            "販管費率 (%)",
            min_value=15.0, max_value=40.0,
            value=sga_default,
            step=0.5,
            format="%.1f%%"
        ) / 100

        st.markdown('<div class="sidebar-header">WACC（加重平均資本コスト）</div>', unsafe_allow_html=True)
        
        # リスクフリーレート（日本10年国債利回り）
        # 自動取得機能あり（デフォルト値を使用）
        rfr_default = 0.87  # 2024年時点の目安
        rfr, rfr_status = get_risk_free_rate()
        if rfr is not None:
            rfr_default = rfr * 100
        
        risk_free_rate = st.slider(
            "リスクフリーレート（Rf） (%)",
            min_value=0.0, max_value=5.0,
            value=rfr_default,
            step=0.01,
            format="%.2f%%",
            help="無リスク資産の期待収益率。通常、日本10年国債利回りを使用。財務省HPで最新値を確認できます。"
        ) / 100

        unlevered_beta = st.slider(
            "アンレバードβ（Unlevered Beta）",
            min_value=0.5, max_value=1.5,
            value=0.90,
            step=0.01,
            format="%.2f",
            help="財務レバレッジの影響を除いた事業リスクを示す指標。類似上場企業のβを平均し、負債比率の影響を除去して算出。業界平均値を参照。"
        )

        target_de_ratio = st.slider(
            "目標D/Eレシオ（Debt/Equity Ratio）",
            min_value=0.0, max_value=2.0,
            value=0.50,
            step=0.05,
            format="%.2f",
            help="目標とする負債と株主資本の比率。業界平均や会社の財務方針に基づいて設定。レバードβの計算とWACCの資本構成比率に使用。"
        )

        equity_risk_premium = st.slider(
            "株式リスクプレミアム（ERP） (%)",
            min_value=4.0, max_value=8.0,
            value=6.0,
            step=0.1,
            format="%.1f%%",
            help="株式投資に対して投資家が要求する超過リターン。市場全体のリスクに対する対価。日本市場では通常5-7%程度。"
        ) / 100

        size_premium = st.slider(
            "サイズプレミアム (%)",
            min_value=0.0, max_value=4.0,
            value=0.0,
            step=0.01,
            format="%.2f%%",
            help="小型株は大型株より高いリターンを示す傾向があるため、時価総額に応じて追加するプレミアム。Ibbotson等のデータを参照。"
        ) / 100

        # 負債コスト（EDINETデータから自動計算可能）
        calculated_cod = st.session_state.get("calculated_cost_of_debt")
        cod_default = calculated_cod * 100 if calculated_cod is not None else 0.80
        cod_default = min(max(cod_default, 0.0), 5.0)  # 範囲内にクリップ
        
        cod_help = "企業が負債（借入金・社債）に対して支払う利率。支払利息÷有利子負債で推計可能。税効果を考慮して税後コストをWACCに使用。"
        if calculated_cod is not None:
            cod_status = st.session_state.get("cost_of_debt_status", "")
            cod_help = f"{cod_help}（{cod_status}）"
        
        cost_of_debt = st.slider(
            "負債コスト（税前）(%)",
            min_value=0.0, max_value=5.0,
            value=cod_default,
            step=0.05,
            format="%.2f%%",
            help=cod_help
        ) / 100

        st.markdown('<div class="sidebar-header">ターミナルバリュー</div>', unsafe_allow_html=True)

        terminal_growth_rate = st.slider(
            "永久成長率 (%)",
            min_value=-1.0, max_value=3.0,
            value=0.0,
            step=0.1,
            format="%.1f%%",
            help="予測期間以降、FCFが永続的に成長する率。通常、長期インフレ率や名目GDP成長率以下に設定。日本では0-1%が保守的。"
        ) / 100

        exit_ebitda_multiple = st.slider(
            "Exit EBITDA Multiple (x)",
            min_value=3.0, max_value=15.0,
            value=8.0,
            step=0.5,
            format="%.1fx",
            help="予測期間終了時点でのEV/EBITDA倍率。類似企業や過去M&A事例の倍率を参考に設定。永久成長率法との整合性をクロスチェック。"
        )

    return {
        "scenario": scenario,
        "base_year": base_year,
        "revenue_growth": revenue_growth,
        "cogs_ratio": cogs_ratio,
        "sga_ratio": sga_ratio,
        "risk_free_rate": risk_free_rate,
        "unlevered_beta": unlevered_beta,
        "target_de_ratio": target_de_ratio,
        "equity_risk_premium": equity_risk_premium,
        "size_premium": size_premium,
        "cost_of_debt": cost_of_debt,
        "terminal_growth_rate": terminal_growth_rate,
        "exit_ebitda_multiple": exit_ebitda_multiple,
        "historical_data": historical_data
    }


def render_value_cards(model: DCFModel):
    """バリューカード表示"""
    summary = model.summary()
    perp = summary["valuation"]["perpetuity"]
    exit_m = summary["valuation"]["exit_multiple"]

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.markdown(f"""
        <div class="value-card">
            <div class="value-card-title">企業価値（EV）</div>
            <div class="value-card-value">{format_currency_oku(perp['enterprise_value'])}</div>
            <div class="value-card-sub">永久成長率法</div>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown(f"""
        <div class="value-card">
            <div class="value-card-title">株式価値</div>
            <div class="value-card-value">{format_currency_oku(perp['equity_value'])}</div>
            <div class="value-card-sub">EV - Net Debt - 非支配持分</div>
        </div>
        """, unsafe_allow_html=True)

    with col3:
        st.markdown(f"""
        <div class="value-card">
            <div class="value-card-title">理論株価</div>
            <div class="value-card-value">{format_stock_price(perp['price_per_share'])}</div>
            <div class="value-card-sub">現在株価: {format_stock_price(model.historical.current_stock_price)}</div>
        </div>
        """, unsafe_allow_html=True)

    with col4:
        upside = (perp['price_per_share'] / model.historical.current_stock_price - 1) * 100
        upside_class = "upside-positive" if upside >= 0 else "upside-negative"
        st.markdown(f"""
        <div class="value-card">
            <div class="value-card-title">アップサイド</div>
            <div class="value-card-value {upside_class}">{upside:+.1f}%</div>
            <div class="value-card-sub">理論株価 vs 現在株価</div>
        </div>
        """, unsafe_allow_html=True)
    
    # 市場ベースEV比較
    render_market_comparison(model, perp)


def render_market_comparison(model: DCFModel, dcf_perp: dict):
    """市場ベースEVとの比較表示"""
    # 市場データ
    market_cap = model.market_cap
    market_ev = model.market_ev
    dcf_ev = dcf_perp['enterprise_value']
    dcf_equity = dcf_perp['equity_value']
    
    # 乖離率
    ev_diff = (dcf_ev / market_ev - 1) * 100 if market_ev > 0 else 0
    equity_diff = (dcf_equity / market_cap - 1) * 100 if market_cap > 0 else 0
    
    st.markdown("""
    <div style="background: #f8f9fa; border-radius: 8px; padding: 0.8rem 1rem; margin-top: 0.5rem;">
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown(f"""
        <div style="text-align: center;">
            <div style="font-size: 0.75rem; color: #666;">市場ベースEV</div>
            <div style="font-size: 1.1rem; font-weight: bold; color: #1e3a5f;">{format_currency_oku(market_ev)}</div>
            <div style="font-size: 0.7rem; color: #888;">時価総額 + Net Debt + 非支配持分</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div style="text-align: center;">
            <div style="font-size: 0.75rem; color: #666;">時価総額</div>
            <div style="font-size: 1.1rem; font-weight: bold; color: #1e3a5f;">{format_currency_oku(market_cap)}</div>
            <div style="font-size: 0.7rem; color: #888;">株価 × 発行済株式数</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        # DCF vs 市場の乖離
        ev_color = "#2e7d32" if ev_diff > 0 else "#c62828"
        equity_color = "#2e7d32" if equity_diff > 0 else "#c62828"
        
        # 状況に応じた説明文
        if equity_diff > 0:
            status_text = "DCF > 市場 → 割安示唆"
        elif equity_diff < 0:
            status_text = "DCF < 市場 → 割高示唆"
        else:
            status_text = "DCF ≈ 市場（適正水準）"
        
        st.markdown(f"""
        <div style="text-align: center;">
            <div style="font-size: 0.75rem; color: #666;">DCF vs 市場（乖離率）</div>
            <div style="font-size: 0.85rem;">
                <span style="color: #666;">EV: </span>
                <span style="font-weight: bold; color: {ev_color};">{ev_diff:+.1f}%</span>
                <span style="color: #666; margin-left: 0.5rem;">株式: </span>
                <span style="font-weight: bold; color: {equity_color};">{equity_diff:+.1f}%</span>
            </div>
            <div style="font-size: 0.7rem; color: #888;">{status_text}</div>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("</div>", unsafe_allow_html=True)


def render_wacc_summary(model: DCFModel):
    """WACC構成要素サマリー"""
    st.markdown('<div class="section-header">WACC構成</div>', unsafe_allow_html=True)
    
    with st.expander("💡 WACCとは？", expanded=False):
        st.markdown("""
        **WACC（Weighted Average Cost of Capital）**は、企業が資金調達するための加重平均コストです。
        
        **計算式**:
        ```
        WACC = CoE × E/(D+E) + CoD × (1-税率) × D/(D+E)
        ```
        
        | 項目 | 説明 |
        |------|------|
        | **株主資本コスト（CoE）** | 株主が期待するリターン。CAPM: Rf + β × ERP + サイズP |
        | **レバードβ** | 財務レバレッジを反映したβ。アンレバードβ × (1 + D/E × (1-税率)) |
        | **負債コスト（税後）** | 借入金利 × (1-税率)。利息の節税効果を反映 |
        | **D/E** | 負債/株主資本比率。資本構成を示す |
        
        WACCはDCFの**割引率**として使用され、将来キャッシュフローを現在価値に換算します。
        """)

    col1, col2, col3 = st.columns(3)

    # 実効税率（日本の法人税率）
    TAX_RATE = 0.30

    with col1:
        st.metric("株主資本コスト（CoE）", format_percent(model.cost_of_equity))
        st.caption(f"レバードβ: {model.levered_beta:.2f}")

    with col2:
        after_tax_cod = model.assumptions.cost_of_debt * (1 - TAX_RATE)
        st.metric("負債コスト（税後）", format_percent(after_tax_cod))
        st.caption(f"税前: {format_percent(model.assumptions.cost_of_debt)}")

    with col3:
        st.metric("WACC", format_percent(model.wacc))
        st.caption(f"D/E: {model.assumptions.target_de_ratio:.2f}")


def render_valuation_bridge(model: DCFModel):
    """バリュエーションブリッジ"""
    st.markdown('<div class="section-header">バリュエーションブリッジ</div>', unsafe_allow_html=True)
    
    with st.expander("💡 バリュエーションブリッジとは？", expanded=False):
        st.markdown("""
        **企業価値（EV）から株式価値への変換過程**を視覚化したチャートです。
        
        | 項目 | 説明 |
        |------|------|
        | **FCF現在価値** | 予測期間中のフリーキャッシュフローをWACCで割り引いた合計値 |
        | **TV現在価値** | ターミナルバリュー（予測期間以降の価値）の現在価値 |
        | **企業価値（EV）** | FCF現在価値 + TV現在価値。事業全体の価値 |
        | **ネットデット** | 有利子負債 - 現金。債権者への返済が必要な金額 |
        | **非支配持分** | 連結子会社のうち親会社以外の株主に帰属する部分 |
        | **株式価値** | EV - ネットデット - 非支配持分。株主に帰属する価値 |
        """)

    summary = model.summary()
    perp = summary["valuation"]["perpetuity"]

    # ウォーターフォールチャート
    fig = go.Figure(go.Waterfall(
        name="Valuation Bridge",
        orientation="v",
        measure=["relative", "relative", "total", "relative", "relative", "total"],
        x=["FCF現在価値", "TV現在価値", "企業価値(EV)", "ネットデット", "非支配持分", "株式価値"],
        textposition="outside",
        text=[
            f"{perp['sum_pv_fcf']:,.0f}",
            f"{perp['pv_terminal_value']:,.0f}",
            f"{perp['enterprise_value']:,.0f}",
            f"{-perp['net_debt']:,.0f}",
            f"{-model.historical.minority_interest:,.0f}",
            f"{perp['equity_value']:,.0f}"
        ],
        y=[
            perp['sum_pv_fcf'],
            perp['pv_terminal_value'],
            None,
            -perp['net_debt'],
            -model.historical.minority_interest,
            None
        ],
        connector={"line": {"color": "#1e3a5f"}},
        increasing={"marker": {"color": "#1e3a5f"}},
        decreasing={"marker": {"color": "#c9a962"}},
        totals={"marker": {"color": "#2d5a8a"}}
    ))

    fig.update_layout(
        title="バリュエーションブリッジ",
        showlegend=False,
        height=450,
        font=dict(family="Arial", size=12),
        margin=dict(t=60, b=60, l=80, r=40),
        yaxis=dict(
            title="百万円",
            tickformat=",",
            gridcolor="#e0e0e0"
        ),
        xaxis=dict(title=None),
        plot_bgcolor="white",
        paper_bgcolor="white"
    )

    st.plotly_chart(fig, use_container_width=True)


def render_fcf_projection(model: DCFModel):
    """FCF予測テーブル"""
    st.markdown('<div class="section-header">FCF予測（百万円）</div>', unsafe_allow_html=True)
    
    with st.expander("💡 FCF予測の項目説明", expanded=False):
        st.markdown("""
        **フリーキャッシュフロー（FCF）**は、企業が自由に使えるキャッシュを示します。
        
        | 項目 | 説明 | 計算式 |
        |------|------|--------|
        | **売上収益** | 予測売上高 | 前期 × (1 + 成長率) |
        | **EBITDA** | 償却前営業利益 | 売上 - 原価 - 販管費 |
        | **減価償却** | 有形固定資産の費用配分 | PPE × 償却率 |
        | **EBIT** | 営業利益 | EBITDA - 減価償却 |
        | **NOPAT** | 税引後営業利益 | EBIT × (1 - 税率) |
        | **設備投資** | 固定資産への投資 | 売上 × CapEx率 |
        | **Δ運転資本** | 運転資本の増減 | 売上成長に伴う追加投資 |
        | **FCF** | フリーキャッシュフロー | NOPAT + 減価償却 - 設備投資 - Δ運転資本 |
        | **割引係数** | 現在価値への換算係数 | 1 / (1 + WACC)^n |
        | **FCF現在価値** | 各年FCFの現在価値 | FCF × 割引係数 |
        """)

    df = model.projection_table()

    # フォーマット適用（カンマ区切り）
    format_rows = {
        "売上収益": lambda x: f"{x:,.0f}",
        "EBITDA": lambda x: f"{x:,.0f}",
        "減価償却": lambda x: f"{x:,.0f}",
        "EBIT": lambda x: f"{x:,.0f}",
        "NOPAT": lambda x: f"{x:,.0f}",
        "設備投資": lambda x: f"{x:,.0f}",
        "Δ運転資本": lambda x: f"{x:,.0f}",
        "FCF": lambda x: f"{x:,.0f}",
        "割引係数": lambda x: f"{x:.4f}",
        "FCF現在価値": lambda x: f"{x:,.0f}"
    }

    df_display = df.copy()
    for row_name, fmt_func in format_rows.items():
        if row_name in df_display.index:
            df_display.loc[row_name] = df_display.loc[row_name].apply(fmt_func)

    # 数値を右寄せにしたHTMLテーブルを表示
    styled_html = df_display.style.set_properties(**{
        'text-align': 'right',
        'font-family': 'Arial, sans-serif',
        'font-size': '0.85rem'
    }).set_table_styles([
        {'selector': 'th', 'props': [('text-align', 'center'), ('font-weight', 'bold')]},
        {'selector': 'th.row_heading', 'props': [('text-align', 'left')]}
    ]).to_html()

    st.markdown(styled_html, unsafe_allow_html=True)

    # FCFグラフ（億円表示）
    fcf_data = model.projection_table().loc["FCF"]
    fcf_in_oku = fcf_data.values / 100  # 百万円 → 億円に変換
    
    fig = go.Figure()

    fig.add_trace(go.Bar(
        x=fcf_data.index,
        y=fcf_in_oku,
        marker_color="#1e3a5f",
        text=[f"{v:,.0f}" for v in fcf_in_oku],
        textposition="outside"
    ))

    fig.update_layout(
        title="FCF推移",
        height=350,
        font=dict(family="Arial", size=12),
        margin=dict(t=50, b=40, l=60, r=40),  # 上部マージン追加
        yaxis=dict(
            title="FCF（億円）",
            tickformat=",",
            gridcolor="#e0e0e0",
            range=[0, max(fcf_in_oku) * 1.2]  # Y軸の範囲を広げる
        ),
        xaxis=dict(title=None),
        plot_bgcolor="white",
        paper_bgcolor="white",
        showlegend=False
    )

    st.plotly_chart(fig, use_container_width=True)


def render_sensitivity_analysis(model: DCFModel):
    """感応度分析"""
    st.markdown('<div class="section-header">感応度分析</div>', unsafe_allow_html=True)
    
    with st.expander("💡 感応度分析とは？", expanded=False):
        st.markdown("""
        **主要な前提条件が変化した場合の株価への影響**を可視化する分析手法です。
        
        | 分析 | 変数 | 目的 |
        |------|------|------|
        | **永久成長率法** | WACC × 永久成長率 | 割引率と成長率の変化に対する感応度を確認 |
        | **Exit Multiple法** | WACC × Exit EBITDA倍率 | 割引率と出口倍率の変化に対する感応度を確認 |
        
        **読み方**:
        - 緑色 → 株価上昇（アップサイド）
        - 赤色 → 株価下落（ダウンサイド）
        - 現在の前提に近い値を中心に、上下のレンジを確認
        
        ⚠️ DCF評価額は前提条件に大きく依存するため、感応度分析で**バリュエーションレンジ**を把握することが重要です。
        """)

    tab1, tab2 = st.tabs(["永久成長率法", "Exit Multiple法"])

    with tab1:
        df_perp = model.sensitivity_analysis("perpetuity")

        # ヒートマップ
        fig = go.Figure(data=go.Heatmap(
            z=df_perp.values,
            x=df_perp.columns.tolist(),
            y=df_perp.index.tolist(),
            colorscale=[
                [0, "#c62828"],
                [0.5, "#ffffff"],
                [1, "#2e7d32"]
            ],
            text=[[f"{v:,.0f}" for v in row] for row in df_perp.values],
            texttemplate="%{text}",
            textfont={"size": 11, "family": "Arial"},
            hovertemplate="WACC: %{x}<br>永久成長率: %{y}<br>理論株価: %{text}<extra></extra>"
        ))

        fig.update_layout(
            title="理論株価（円）- WACC vs 永久成長率",
            height=400,
            font=dict(family="Arial", size=12),
            xaxis=dict(title="WACC", side="bottom"),
            yaxis=dict(title="永久成長率", autorange="reversed"),
            paper_bgcolor="white"
        )

        st.plotly_chart(fig, use_container_width=True)

        # テーブル表示
        with st.expander("詳細テーブル"):
            df_formatted = df_perp.map(lambda x: f"{x:,.0f}" if pd.notna(x) else "-")
            st.dataframe(df_formatted, use_container_width=True)

    with tab2:
        df_exit = model.sensitivity_analysis("exit_multiple")

        # ヒートマップ
        fig = go.Figure(data=go.Heatmap(
            z=df_exit.values,
            x=df_exit.columns.tolist(),
            y=df_exit.index.tolist(),
            colorscale=[
                [0, "#c62828"],
                [0.5, "#ffffff"],
                [1, "#2e7d32"]
            ],
            text=[[f"{v:,.0f}" for v in row] for row in df_exit.values],
            texttemplate="%{text}",
            textfont={"size": 11, "family": "Arial"},
            hovertemplate="WACC: %{x}<br>Exit Multiple: %{y}<br>理論株価: %{text}<extra></extra>"
        ))

        fig.update_layout(
            title="理論株価（円）- WACC vs Exit Multiple",
            height=400,
            font=dict(family="Arial", size=12),
            xaxis=dict(title="WACC", side="bottom"),
            yaxis=dict(title="Exit Multiple", autorange="reversed"),
            paper_bgcolor="white"
        )

        st.plotly_chart(fig, use_container_width=True)

        # テーブル表示
        with st.expander("詳細テーブル"):
            df_formatted = df_exit.map(lambda x: f"{x:,.0f}" if pd.notna(x) else "-")
            st.dataframe(df_formatted, use_container_width=True)


def render_scenario_comparison(inputs: dict):
    """複数シナリオ比較"""
    st.markdown('<div class="section-header">シナリオ比較</div>', unsafe_allow_html=True)
    
    with st.expander("💡 シナリオ比較とは？", expanded=False):
        st.markdown("""
        **異なる成長率前提での理論株価レンジ**を比較します。
        
        | シナリオ | 特徴 |
        |---------|------|
        | **強気（Bull）** | 楽観的な成長率前提。市場環境が良好な場合 |
        | **ベース（Base）** | 最も蓋然性の高い前提。標準ケース |
        | **弱気（Bear）** | 保守的な成長率前提。リスクを考慮 |
        
        投資判断では、**ベースケースを中心に、Bull〜Bearのレンジ**を参考にします。
        """)
    
    # 現在の成長率を基準に±5%でシナリオ作成
    base_growth = inputs["revenue_growth"]
    scenarios = {
        Scenario.BULL: min(base_growth + 0.05, 0.30),
        Scenario.BASE: base_growth,
        Scenario.BEAR: max(base_growth - 0.05, 0.0),
    }
    
    # 各シナリオでモデル計算
    results = {}
    for scenario, growth in scenarios.items():
        scenario_inputs = inputs.copy()
        scenario_inputs["scenario"] = scenario
        scenario_inputs["revenue_growth"] = growth
        model = create_model_from_inputs(**scenario_inputs)
        summary = model.summary()
        results[scenario] = {
            "growth": growth,
            "ev": summary["valuation"]["perpetuity"]["enterprise_value"],
            "equity": summary["valuation"]["perpetuity"]["equity_value"],
            "price": summary["valuation"]["perpetuity"]["price_per_share"],
            "wacc": summary["wacc"]["wacc"],
        }
    
    # カード表示
    col1, col2, col3 = st.columns(3)
    
    scenario_colors = {
        Scenario.BULL: "#2e7d32",
        Scenario.BASE: "#1e3a5f",
        Scenario.BEAR: "#c62828"
    }
    
    for col, scenario in zip([col1, col2, col3], [Scenario.BEAR, Scenario.BASE, Scenario.BULL]):
        r = results[scenario]
        with col:
            st.markdown(f"""
            <div style="background: white; border-radius: 10px; padding: 1rem; 
                        border-left: 4px solid {scenario_colors[scenario]}; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <div style="color: {scenario_colors[scenario]}; font-weight: bold; font-size: 0.9rem;">
                    {scenario.value}シナリオ
                </div>
                <div style="font-size: 0.8rem; color: #666; margin: 0.3rem 0;">
                    成長率: {r['growth']:.1%}
                </div>
                <div style="font-size: 1.5rem; font-weight: bold; color: #1e3a5f;">
                    {r['price']:,.0f}円
                </div>
                <div style="font-size: 0.75rem; color: #888;">
                    EV: {r['ev']/100:,.0f}億円 / 株式価値: {r['equity']/100:,.0f}億円
                </div>
            </div>
            """, unsafe_allow_html=True)
    
    # 理論株価レンジ表示
    prices = [results[s]["price"] for s in [Scenario.BEAR, Scenario.BASE, Scenario.BULL]]
    current_price = inputs["historical_data"].current_stock_price
    
    st.markdown(f"""
    <div style="background: #f8f9fa; border-radius: 8px; padding: 1rem; margin-top: 1rem; text-align: center;">
        <span style="font-size: 0.85rem; color: #666;">理論株価レンジ: </span>
        <span style="font-size: 1.1rem; font-weight: bold; color: #c62828;">{prices[0]:,.0f}円</span>
        <span style="color: #666;"> 〜 </span>
        <span style="font-size: 1.1rem; font-weight: bold; color: #1e3a5f;">{prices[1]:,.0f}円</span>
        <span style="color: #666;"> 〜 </span>
        <span style="font-size: 1.1rem; font-weight: bold; color: #2e7d32;">{prices[2]:,.0f}円</span>
        <span style="font-size: 0.85rem; color: #666; margin-left: 1rem;">（現在株価: {current_price:,.0f}円）</span>
    </div>
    """, unsafe_allow_html=True)


def render_multiple_valuation(model: DCFModel, inputs: dict):
    """マルチプル法バリュエーション"""
    st.markdown('<div class="section-header">マルチプル法バリュエーション</div>', unsafe_allow_html=True)
    
    with st.expander("💡 マルチプル法とは？", expanded=False):
        st.markdown("""
        **類似企業や市場の倍率を使った簡易バリュエーション**手法です。
        
        | 手法 | 計算式 | 特徴 |
        |------|--------|------|
        | **EV/EBITDA法** | EV = EBITDA × 倍率 | 資本構成の影響を排除。M&Aで頻用 |
        | **PER法** | 株式価値 = 純利益 × PER | 最も一般的。利益ベースの評価 |
        
        DCF法と併用することで、**バリュエーションの妥当性をクロスチェック**できます。
        """)
    
    # データの有無を確認
    has_ebitda = model.historical.ebitda > 0
    has_net_income = model.historical.net_income > 0
    
    if not has_ebitda and not has_net_income:
        st.warning("マルチプル法の計算にはEBITDAまたは当期純利益のデータが必要です。CSVから営業利益・減価償却費・当期純利益を抽出できませんでした。")
        return
    
    # 倍率入力
    col1, col2 = st.columns(2)
    
    with col1:
        ev_ebitda_input = st.number_input(
            "EV/EBITDA倍率 (x)",
            min_value=1.0,
            max_value=30.0,
            value=8.0,
            step=0.5,
            help="類似企業のEV/EBITDA倍率を入力",
            disabled=not has_ebitda
        )
    
    with col2:
        per_input = st.number_input(
            "PER倍率 (x)",
            min_value=1.0,
            max_value=50.0,
            value=15.0,
            step=0.5,
            help="類似企業のPER倍率を入力",
            disabled=not has_net_income
        )
    
    # マルチプル計算
    mult_result = model.multiple_valuation(
        ev_ebitda_multiple=ev_ebitda_input if has_ebitda else None,
        per_multiple=per_input if has_net_income else None
    )
    
    # DCF結果取得
    dcf_summary = model.summary()
    dcf_price = dcf_summary["valuation"]["perpetuity"]["price_per_share"]
    current_price = model.historical.current_stock_price
    
    # 結果表示
    st.markdown("### 理論株価比較")
    
    # データがある手法のみリストに追加
    methods = [("DCF法（永久成長率）", dcf_price, "#1e3a5f")]
    if has_ebitda:
        methods.append(("EV/EBITDA法", mult_result["ev_ebitda"]["price_per_share"], "#2d5a8a"))
    if has_net_income:
        methods.append(("PER法", mult_result["per"]["price_per_share"], "#4a7c9b"))
    methods.append(("現在株価", current_price, "#c9a962"))
    
    # 横棒グラフ（フットボールチャート風）
    fig = go.Figure()
    
    for i, (name, price, color) in enumerate(methods):
        fig.add_trace(go.Bar(
            y=[name],
            x=[price],
            orientation='h',
            marker_color=color,
            text=f"{price:,.0f}円",
            textposition='outside',
            textfont=dict(size=12),
            name=name
        ))
    
    fig.update_layout(
        title="バリュエーション手法別 理論株価",
        height=250 + len(methods) * 30,
        showlegend=False,
        xaxis=dict(title="株価（円）", tickformat=","),
        yaxis=dict(title=None, autorange="reversed"),
        plot_bgcolor="white",
        paper_bgcolor="white",
        margin=dict(l=150, r=80, t=50, b=40),
        font=dict(family="Arial", size=12)
    )
    
    # 現在株価の縦線
    fig.add_vline(
        x=current_price,
        line_dash="dash",
        line_color="#c9a962",
        annotation_text=f"現在株価: {current_price:,.0f}円",
        annotation_position="top"
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # 詳細テーブル
    with st.expander("詳細データ"):
        # データがある手法のみ表示
        rows_method = ["DCF法"]
        rows_multiple = [f"WACC: {dcf_summary['wacc']['wacc']:.2%}, 永久成長率: {model.assumptions.terminal_growth_rate:.1%}"]
        rows_ev = [f"{dcf_summary['valuation']['perpetuity']['enterprise_value']/100:,.0f}"]
        rows_equity = [f"{dcf_summary['valuation']['perpetuity']['equity_value']/100:,.0f}"]
        rows_price = [f"{dcf_price:,.0f}"]
        rows_upside = [f"{(dcf_price/current_price - 1)*100:+.1f}%"]
        
        if has_ebitda:
            rows_method.append("EV/EBITDA法")
            rows_multiple.append(f"{ev_ebitda_input:.1f}x")
            rows_ev.append(f"{mult_result['ev_ebitda']['enterprise_value']/100:,.0f}")
            rows_equity.append(f"{mult_result['ev_ebitda']['equity_value']/100:,.0f}")
            rows_price.append(f"{mult_result['ev_ebitda']['price_per_share']:,.0f}")
            rows_upside.append(f"{(mult_result['ev_ebitda']['price_per_share']/current_price - 1)*100:+.1f}%")
        
        if has_net_income:
            rows_method.append("PER法")
            rows_multiple.append(f"{per_input:.1f}x")
            rows_ev.append("—")
            rows_equity.append(f"{mult_result['per']['equity_value']/100:,.0f}")
            rows_price.append(f"{mult_result['per']['price_per_share']:,.0f}")
            rows_upside.append(f"{(mult_result['per']['price_per_share']/current_price - 1)*100:+.1f}%")
        
        detail_data = {
            "手法": rows_method,
            "倍率/前提": rows_multiple,
            "企業価値（億円）": rows_ev,
            "株式価値（億円）": rows_equity,
            "理論株価（円）": rows_price,
            "現在株価比": rows_upside
        }
        st.dataframe(pd.DataFrame(detail_data), hide_index=True, use_container_width=True)
        
        # 市場データ
        st.markdown("**参考: 市場データ**")
        st.write(f"- 時価総額: {model.market_cap/100:,.0f}億円")
        st.write(f"- 市場ベースEV: {model.market_ev/100:,.0f}億円")
        if has_ebitda:
            st.write(f"- 実績EV/EBITDA: {model.market_ev_ebitda:.1f}x")
        if has_net_income:
            st.write(f"- 実績PER: {model.market_per:.1f}x")


def render_cross_check(model: DCFModel):
    """クロスチェック"""
    st.markdown('<div class="section-header">クロスチェック</div>', unsafe_allow_html=True)
    
    with st.expander("💡 クロスチェックとは？", expanded=False):
        st.markdown("""
        **2つのターミナルバリュー算定方法の整合性**を確認するための指標です。
        
        | 指標 | 説明 | 目安 |
        |------|------|------|
        | **インプライドEV/EBITDA** | 永久成長率法から逆算されるExit Multiple。設定したExit Multipleと大きく乖離していないか確認 | 業界平均±2x程度 |
        | **インプライド永久成長率** | Exit Multiple法から逆算される永久成長率。非現実的に高い/低い値でないか確認 | -1%〜3%程度 |
        | **TV/EV比率** | ターミナルバリューが企業価値に占める割合。高すぎる場合は予測期間のCFへの依存度が低い | 50-80%程度 |
        
        ⚠️ **注意**: TV/EV比率が80%を超える場合、ターミナルバリューへの依存度が高く、前提の変化に対して敏感になります。
        """)

    summary = model.summary()
    perp = summary["valuation"]["perpetuity"]
    exit_m = summary["valuation"]["exit_multiple"]

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**永久成長率法**")
        st.write(f"- インプライドEV/EBITDA: **{perp['implied_exit_multiple']:.1f}x**")
        st.write(f"- TV/EV比率: **{perp['tv_percentage']:.1%}**")

    with col2:
        st.markdown("**Exit Multiple法**")
        st.write(f"- インプライド永久成長率: **{exit_m['implied_growth_rate']:.2%}**")
        st.write(f"- TV/EV比率: **{exit_m['tv_percentage']:.1%}**")


def main():
    """メイン関数"""
    # ヘッダー
    render_header()

    # サイドバーから入力値取得
    inputs = render_sidebar()

    # データがアップロードされていない場合は案内を表示
    if inputs["historical_data"] is None:
        st.markdown("---")
        st.info(
            "👈 **サイドバーからデータをアップロードしてください**\n\n"
            "- **EDINET CSV**: 上場企業の有価証券報告書CSVファイル\n"
            "- **Excelテンプレート**: 非上場企業向けテンプレート\n\n"
            "データをアップロード後、「データを適用」ボタンをクリックするとバリュエーション結果が表示されます。"
        )
        
        # フッター
        st.markdown("---")
        st.markdown(
            '<p style="text-align: center; color: #888; font-size: 0.8rem;">'
            'DCF Valuation Model | Built with Streamlit</p>',
            unsafe_allow_html=True
        )
        return

    # モデル作成
    model = create_model_from_inputs(**inputs)

    # シナリオ表示
    scenario_colors = {
        Scenario.BULL: "#2e7d32",
        Scenario.BASE: "#1e3a5f",
        Scenario.BEAR: "#c62828"
    }
    st.markdown(
        f'<span style="background-color: {scenario_colors[inputs["scenario"]]}; '
        f'color: white; padding: 0.3rem 0.8rem; border-radius: 15px; font-size: 0.9rem;">'
        f'{inputs["scenario"].value}シナリオ</span>',
        unsafe_allow_html=True
    )

    st.write("")

    # バリューカード
    render_value_cards(model)

    # WACC構成
    render_wacc_summary(model)

    # 2カラムレイアウト
    col1, col2 = st.columns([3, 2])

    with col1:
        # バリュエーションブリッジ
        render_valuation_bridge(model)

    with col2:
        # クロスチェック
        render_cross_check(model)

    # FCF予測
    render_fcf_projection(model)

    # 感応度分析
    render_sensitivity_analysis(model)
    
    # シナリオ比較
    render_scenario_comparison(inputs)
    
    # マルチプル法バリュエーション
    render_multiple_valuation(model, inputs)

    # エクスポートセクション
    st.markdown("---")
    st.markdown('<div class="section-header">分析結果のエクスポート</div>', unsafe_allow_html=True)
    
    # ファイル形式選択
    export_format = st.radio(
        "出力形式を選択",
        options=["Excel", "PDF"],
        horizontal=True,
        help="Excel: 詳細データ・編集可能 / PDF: レポート形式・印刷向け"
    )
    
    # ファイル名生成（日付付き）
    today_str = datetime.now().strftime("%Y%m%d")
    company_name = st.session_state.get("company_name", "")
    base_filename = f"DCF_{company_name}_{today_str}" if company_name else f"DCF_分析結果_{today_str}"
    
    col1, col2 = st.columns(2)
    
    with col1:
        if export_format == "Excel":
            excel_file = create_analysis_excel(model, inputs)
            st.download_button(
                label="📥 Excelでダウンロード",
                data=excel_file,
                file_name=f"{base_filename}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="サマリー、FCF予測、感応度分析を含むExcelファイル",
                use_container_width=True
            )
        else:
            pdf_file = create_analysis_pdf(model, inputs)
            st.download_button(
                label="📥 PDFでダウンロード",
                data=pdf_file,
                file_name=f"{base_filename}.pdf",
                mime="application/pdf",
                help="レポート形式のPDFファイル（印刷向け）",
                use_container_width=True
            )
    
    with col2:
        # もう一方の形式もダウンロード可能に
        if export_format == "Excel":
            pdf_file = create_analysis_pdf(model, inputs)
            st.download_button(
                label="📄 PDFもダウンロード",
                data=pdf_file,
                file_name=f"{base_filename}.pdf",
                mime="application/pdf",
                help="レポート形式のPDFファイル",
                use_container_width=True
            )
        else:
            excel_file = create_analysis_excel(model, inputs)
            st.download_button(
                label="📊 Excelもダウンロード",
                data=excel_file,
                file_name=f"{base_filename}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="詳細データを含むExcelファイル",
                use_container_width=True
            )

    # フッター
    company_name = st.session_state.get("company_name", "")
    footer_text = f"DCF Valuation Model - {company_name}" if company_name else "DCF Valuation Model"
    st.markdown("---")
    st.markdown(
        f'<p style="text-align: center; color: #888; font-size: 0.8rem;">'
        f'{footer_text} | Built with Streamlit</p>',
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()
